"""
excel_loader.py
----------------
Loads "Breeding_charts.xlsx" and, for every map (Freshwater, Saltwater,
Magical), builds breeding lookup tables for both FINS and BODIES.

Expected sheet layout in the workbook (for each map "<Map>"):
    "<Map>_FINS"    - fin breeding matrix
    "<Map>_BODIES"  - body breeding matrix

Sheet structure:
    - Row 1 (from column B) = list of species (column headers)
    - Column A (from row 2) = the same list of species (row headers)
    - Cells (row 2..N+1, column B..N+1) = the resulting offspring species
    - The cell's background color encodes the rarity of that specific
      parent combination:
        theme 9 (accent6, green)    -> Common
        theme 4 (accent1, teal/blue) -> Uncommon
        theme 7 (accent4, bright blue) -> treated as Uncommon by agreement
                                           (inconsistent formatting in the
                                           original Excel file, mostly on
                                           self-crosses)
        red RGB (FFFF0000)          -> Rare
    - Below the matrix there are 3 legend rows (Rare / Uncommon / Common);
      those are skipped during parsing since we detect the end of the
      matrix from the length of the species list.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl

RARITY_ORDER = {"Common": 0, "Uncommon": 1, "Rare": 2}

MAP_NAMES = ["Freshwater", "Saltwater", "Magical"]


def _cell_rarity(cell) -> str:
    """Determine rarity from the cell's background fill color."""
    fg = cell.fill.fgColor
    if fg is None:
        return "Common"

    if fg.type == "rgb":
        rgb = fg.rgb
        if isinstance(rgb, str) and "FF0000" in rgb.upper():
            return "Rare"
        return "Common"

    if fg.type == "theme":
        theme = fg.theme
        if theme == 9:
            return "Common"
        if theme == 4:
            return "Uncommon"
        if theme == 7:
            # Inconsistent/"stray" color in the original file - treated as
            # Uncommon by agreement with the user.
            return "Uncommon"
        return "Common"

    return "Common"


@dataclass
class BreedTable:
    """A single breeding matrix (either for FINS or for BODIES)."""

    species: List[str] = field(default_factory=list)
    # key: frozenset({a, b}), or (a,) for a self-cross
    results: Dict[frozenset, Tuple[str, str]] = field(default_factory=dict)

    def _key(self, a: str, b: str):
        return frozenset((a, b)) if a != b else frozenset((a,))

    def breed(self, a: str, b: str) -> Optional[Tuple[str, str]]:
        """Returns (resulting_species, rarity), or None if the combination
        isn't in the table."""
        return self.results.get(self._key(a, b))


@dataclass
class MagicFish:
    title: str
    body: str
    fin: str


@dataclass
class MapData:
    name: str
    fins: BreedTable
    bodies: BreedTable
    magic_fish: List[MagicFish] = field(default_factory=list)

    @property
    def fin_species(self) -> List[str]:
        return self.fins.species

    @property
    def body_species(self) -> List[str]:
        return self.bodies.species

    def breed_fish(self, fish_a: Tuple[str, str], fish_b: Tuple[str, str]):
        """
        Breeds two fish (body, fin) x (body, fin).
        Returns ((result_body, result_fin), (body_rarity, fin_rarity)) or
        None if either sub-combination isn't in the tables.
        """
        body_a, fin_a = fish_a
        body_b, fin_b = fish_b
        body_res = self.bodies.breed(body_a, body_b)
        fin_res = self.fins.breed(fin_a, fin_b)
        if body_res is None or fin_res is None:
            return None
        (new_body, body_rarity) = body_res
        (new_fin, fin_rarity) = fin_res
        return (new_body, new_fin), (body_rarity, fin_rarity)


def _load_table(ws) -> BreedTable:
    # Read the species list from row 1 (columns B, C, ... until empty)
    species: List[str] = []
    col = 2
    while True:
        val = ws.cell(row=1, column=col).value
        if val is None or str(val).strip() == "":
            break
        species.append(str(val).strip())
        col += 1

    n = len(species)
    table = BreedTable(species=species)

    for i in range(n):
        row_species = species[i]
        for j in range(n):
            col_species = species[j]
            cell = ws.cell(row=2 + i, column=2 + j)
            val = cell.value
            if val is None:
                continue
            rarity = _cell_rarity(cell)
            key = table._key(row_species, col_species)
            # The matrix should be symmetric; if we already have a value,
            # keep the first one we saw.
            if key not in table.results:
                table.results[key] = (str(val).strip(), rarity)

    return table


def _load_magic_fish(ws) -> List[MagicFish]:
    """Parses a "<Map>_MAGICFISH" sheet. Expected columns (with a header row):
    A = Fin, B = Body, C = Magic Fish Title."""
    magic_fish: List[MagicFish] = []
    row = 2
    while True:
        fin = ws.cell(row=row, column=1).value
        body = ws.cell(row=row, column=2).value
        title = ws.cell(row=row, column=3).value
        if fin is None and body is None and title is None:
            break
        if fin is not None and body is not None and title is not None:
            magic_fish.append(
                MagicFish(title=str(title).strip(), body=str(body).strip(), fin=str(fin).strip())
            )
        row += 1
    return magic_fish


def load_all_maps(xlsx_path: str) -> Dict[str, MapData]:
    """Loads every available map from the file. A map is skipped if its
    sheets are missing (e.g. Saltwater/Magical before they're added)."""
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_names = set(wb.sheetnames)

    maps: Dict[str, MapData] = {}
    for map_name in MAP_NAMES:
        fins_sheet = f"{map_name}_FINS"
        bodies_sheet = f"{map_name}_BODIES"
        if fins_sheet in sheet_names and bodies_sheet in sheet_names:
            fins = _load_table(wb[fins_sheet])
            bodies = _load_table(wb[bodies_sheet])

            magic_fish: List[MagicFish] = []
            magic_sheet = f"{map_name}_MAGICFISH"
            if magic_sheet in sheet_names:
                magic_fish = _load_magic_fish(wb[magic_sheet])

            maps[map_name] = MapData(name=map_name, fins=fins, bodies=bodies, magic_fish=magic_fish)

    return maps


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    maps = load_all_maps(os.path.join(here, "data", "Breeding_charts.xlsx"))
    for name, m in maps.items():
        print(name, "fins:", len(m.fin_species), "bodies:", len(m.body_species))
